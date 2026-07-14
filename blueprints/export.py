"""Exportação: planilha .xlsx (espelho) e espelho de ponto em PDF."""
import io
from datetime import date
from tempo import hoje as _hoje
from flask import Blueprint, request, send_file, render_template
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from models import db, Colaborador
from blueprints.auth import admin_req
from servico import calcula_colaborador_mes
from jornada import fmt_hm

bp = Blueprint("export", __name__, url_prefix="/admin/export")

MESES = ["", "Janeiro", "Fevereiro", "Março", "Abril", "Maio", "Junho",
         "Julho", "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro"]
DIAS_ABREV = ["Seg", "Ter", "Qua", "Qui", "Sex", "Sáb", "Dom"]


def _periodo():
    hoje = _hoje()
    return int(request.args.get("ano", hoje.year)), int(request.args.get("mes", hoje.month))


def gera_xlsx(colaboradores, ano, mes):
    wb = Workbook()
    wb.remove(wb.active)
    head_fill = PatternFill("solid", fgColor="0F3D3E")
    head_font = Font(color="FFFFFF", bold=True)
    center = Alignment(horizontal="center")
    thin = Border(*[Side(style="thin", color="DDDDDD")] * 4)

    for c in colaboradores:
        dias, t = calcula_colaborador_mes(c, ano, mes)
        ws = wb.create_sheet(title=c.nome[:28])
        ws.append([f"Espelho de Ponto — {c.nome} — {MESES[mes]}/{ano}"])
        ws["A1"].font = Font(bold=True, size=13)
        ws.append([])
        cab = ["Dia", "Semana", "Trabalhado", "Intervalo", "Atraso",
               "HE 50%", "HE 100%", "Noturno", "Situação"]
        ws.append(cab)
        for col in range(1, len(cab) + 1):
            cell = ws.cell(row=3, column=col)
            cell.fill, cell.font, cell.alignment = head_fill, head_font, center

        for d in dias:
            situacao = []
            if d["falta"]:
                situacao.append("Falta")
            if d["abonado"] and not d["trabalhou"]:
                situacao.append("Abonado")
            if d["feriado"]:
                situacao.append("Feriado")
            if d["aberto"]:
                situacao.append("Jornada aberta")
            ws.append([
                d["dia"].strftime("%d/%m"),
                DIAS_ABREV[d["dia"].weekday()],
                fmt_hm(d["liquido_min"]) if d["liquido_min"] else "",
                fmt_hm(d["intervalo_min"]) if d["intervalo_min"] else "",
                fmt_hm(d["atraso_min"]) if d["atraso_min"] else "",
                fmt_hm(d["he50_min"]) if d["he50_min"] else "",
                fmt_hm(d["he100_min"]) if d["he100_min"] else "",
                fmt_hm(d["noturno_min"]) if d["noturno_min"] else "",
                ", ".join(situacao),
            ])

        ws.append([])
        resumo = [
            ("Horas trabalhadas", fmt_hm(t["liquido_min"])),
            ("HE 50%", fmt_hm(t["he50_min"])),
            ("HE 100%", fmt_hm(t["he100_min"])),
            ("HE total", fmt_hm(t["he_total_min"])),
            ("HE pagável (até 40h)", fmt_hm(t["he_pagavel_min"])),
            ("Excedente p/ banco de horas", fmt_hm(t["banco_min"])),
            ("Adicional noturno", fmt_hm(t["noturno_min"])),
            ("Atrasos", fmt_hm(t["atraso_min"])),
            ("Faltas", str(t["faltas"])),
            ("Faltas abonadas", str(t["faltas_abonadas"])),
            ("Dias trabalhados", str(t["dias_trabalhados"])),
            ("Jornadas em aberto", str(t["dias_aberto"])),
        ]
        for nome, val in resumo:
            ws.append([nome, val])
            ws.cell(row=ws.max_row, column=1).font = Font(bold=True)

        widths = [10, 9, 12, 11, 9, 9, 9, 10, 22]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[chr(64 + i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


@bp.route("/xlsx")
@admin_req
def xlsx():
    ano, mes = _periodo()
    cid = request.args.get("cid")
    if cid:
        colaboradores = [db.session.get(Colaborador, int(cid))]
    else:
        colaboradores = Colaborador.query.filter_by(ativo=True).order_by(Colaborador.nome).all()
    buf = gera_xlsx(colaboradores, ano, mes)
    nome = f"espelho_{MESES[mes].lower()}_{ano}.xlsx"
    return send_file(buf, as_attachment=True, download_name=nome,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@bp.route("/espelho/<int:cid>")
@admin_req
def espelho_pdf(cid):
    from weasyprint import HTML
    ano, mes = _periodo()
    c = db.session.get(Colaborador, cid)
    dias, t = calcula_colaborador_mes(c, ano, mes)
    html = render_template("espelho_pdf.html", colab=c, dias=dias, t=t,
                           ano=ano, mes=mes, mes_nome=MESES[mes], fmt=fmt_hm,
                           dias_abrev=DIAS_ABREV)
    pdf = HTML(string=html).write_pdf()
    nome = f"espelho_{c.nome.replace(' ', '_')}_{MESES[mes].lower()}_{ano}.pdf"
    return send_file(io.BytesIO(pdf), as_attachment=True, download_name=nome,
                     mimetype="application/pdf")
